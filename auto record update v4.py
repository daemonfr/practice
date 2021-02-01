import os
import openpyxl
import glob
from openpyxl import Workbook
import win32com.client
import re
import datetime as dt
import pandas as pd

path = r'C:\Test\Weekly review form'

wb = Workbook()

#return to the home directory
def home_dir():
    os.chdir(path)
    return

#open excel record
def open_review_form():
    global wb
    for filename in glob.glob(os.path.join(path, '*.xlsx')):
        print(filename)    
        wb = openpyxl.open(filename)
    return

#open the intern tab in the excel record
def open_tab(tab_name):
    global ws
    ws = wb[tab_name]
    return

#list out all employee names (including both intern and account manager)
def employee_names(col_name):
    global employee_list
    global ws
    employee_list = []
    for col in ws[col_name]:
        employee_list.append(col.value)
    employee_list.pop(0)
    return employee_list

def timestamp():
    timestamp = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return str(timestamp)

#function to add to the list of names submitted forms
def add_names(iterable, patt, title):
    global dict_of_intern_completed_form
    global dict_of_acc_mgr_completed_form
    global num_of_intern_completed_form
    global num_of_acc_mgr_completed_form
    global num_of_intern
    global num_of_acc_mgr
    
    txt = re.compile('Name \* \\t([A-z][a-z]+ [A-Z][a-z]+ \([A-Z][a-z]+\)|[A-z][a-z]+ [A-Z][a-z]+|[A-z][a-z]+)', re.IGNORECASE)
    pattern = re.compile(patt, re.IGNORECASE)
    
    if title == 'intern':
        if pattern.findall(iterable.Subject) == 'Weekly Intern Review Form':
            num_of_intern_completed_form = num_of_intern_completed_form + 1
            found = False
            date = iterable.SentOn.strftime("%d-%m-%Y")
            if txt.findall(iterable.Body):
                if date not in dict_of_intern_completed_form:
                    dict_of_intern_completed_form[date] = [txt.findall(iterable.Body)[0]]
                else:
                    name = str(txt.findall(iterable.Body)[0])
                    dict_of_intern_completed_form[date].append(name)
                found = True
                print(timestamp() + ': Name added to email without errors.')
            if not found:
                print(timestamp() + ': No name found for email, although subject matches.')
    
    if title == 'acc_mgr':
        if pattern.findall(iterable.Subject) == 'Weekly Report - Account Manager':
            num_of_acc_mgr_completed_form = num_of_acc_mgr_completed_form + 1
            found = False
            date = iterable.SentOn.strftime("%d-%m-%Y")
            if txt.findall(iterable.Body):
                if date not in dict_of_acc_mgr_completed_form:
                    dict_of_acc_mgr_completed_form[date] = [txt.findall(iterable.Body)[0]]
                else:
                    name = str(txt.findall(iterable.Body)[0])
                    dict_of_acc_mgr_completed_form[date].append(name)
                found = True
                print(timestamp() + ': Name added to email without errors.')
            if not found:
                print(timestamp() + ': No name found for email, although subject matches.')
    

def check_inbox(mail_box):
    global dict_of_intern_completed_form
    global dict_of_acc_mgr_completed_form
    outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')
    inbox = outlook.Folders(mail_box).Folders('Inbox')    
    messages = inbox.Items
    
    dict_of_intern_completed_form = {}
    dict_of_acc_mgr_completed_form = {}
    num_of_intern_completed_form = 0
    num_of_acc_mgr_completed_form = 0
    
    txt = re.compile('Name \* \\t([A-z][a-z]+ [A-Z][a-z]+ \([A-Z][a-z]+\)|[A-z][a-z]+ [A-Z][a-z]+|[A-z][a-z]+)', re.IGNORECASE)
    
    search_subject = ['Weekly Intern Review Form', 'Weekly Report - Account Manager']
    
    for h, i in enumerate(search_subject):
        print(timestamp() + ': Looking for ' + str(i) + ' in inbox')
        count = 0
        pattern = re.compile(i, re.IGNORECASE)
        for j, k in enumerate(messages):
            count = j + 1
            print(timestamp() + ': Parsing item ' + str(count))
    
            if pattern.findall(k.Subject):
                print(timestamp() + ': Found a matching subject for email number: ' + str(count))
                if h == 0:
                     
                    if pattern.findall(k.Subject):
                        num_of_intern_completed_form = num_of_intern_completed_form + 1
                        found = False
                        date = k.SentOn.strftime("%d-%m-%Y")
                        
                        if txt.findall(k.Body):
                            if date not in dict_of_intern_completed_form:
                                dict_of_intern_completed_form[date] = [txt.findall(k.Body)[0]]
                            else:
                                name = str(txt.findall(k.Body)[0])
                                dict_of_intern_completed_form[date].append(name)
                            found = True
                            print(timestamp() + ': Name added to email without errors.')
                        if not found:
                            print(timestamp() + ': No name found for email, although subject matches.')
                    
                if h == 1:
                    
                    if pattern.findall(k.Subject):
                        num_of_acc_mgr_completed_form = num_of_acc_mgr_completed_form + 1
                        found = False
                        date = k.SentOn.strftime("%d-%m-%Y")
                        if txt.findall(k.Body):
                            if date not in dict_of_acc_mgr_completed_form:
                                dict_of_acc_mgr_completed_form[date] = [txt.findall(k.Body)[0]]
                            else:
                                name = str(txt.findall(k.Body)[0])
                                dict_of_acc_mgr_completed_form[date].append(name)
                            found = True
                            print(timestamp() + ': Name added to email without errors.')
                        if not found:
                            print(timestamp() + ': No name found for email, although subject matches.')
                            
            else:
                print(timestamp()+ ': No match for item ' + str(count))
        
    if num_of_intern_completed_form == len(dict_of_intern_completed_form):
        print('The number of people completed form matches with the length of the extracted list of people completed the form')
    else:
        print('ERROR: ' + str(num_of_intern_completed_form) + ' completed the form but only ' + str(len(dict_of_intern_completed_form)) + ' people in the dictionary.') 
    if num_of_acc_mgr_completed_form == len(dict_of_acc_mgr_completed_form):
        print('The number of people completed form matches with the length of the extracted list of people completed the form')
    else:
        print('ERROR: ' + str(num_of_acc_mgr_completed_form) + ' completed the form but only ' + str(len(dict_of_acc_mgr_completed_form)) + ' people in the dictionary.')  
        
    
#open record file
home_dir()
open_review_form()

#read the intern 2021 tab
open_tab('intern 2021')
#ws['E3'].value
employee_names('B')

#read the consultant tab
open_tab('consultant')
employee_names('B')

#check Career UK email for form completion
check_inbox('Career UK')



filename = 'results.xlsx'

result_wb = openpyxl.load_workbook(filename)

writer = pd.ExcelWriter(filename, engine='openpyxl')

dict_of_intern_completed_form.keys()

for sheet, frame in dict_of_intern_completed_form.items():
    writer.sheets = dict((ws.title, ws) for ws in result_wb.worksheets) # need this to prevent overwrite
    frame.to_excel(writer, index=False, sheet_name = Sheet1)

writer.save()

result_wb = openpyxl.load_workbook(filename)
for ws in result_wb.worksheets:
   mxrow = ws.max_row
   mxcol = ws.max_column
   tab = openpyxl.worksheet.table.Table(displayName=ws.title, ref="A1:" + ws.cell(mxrow,mxcol).coordinate)
   ws.add_table(tab)

result_wb.save(filename)













