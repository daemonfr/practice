import os
import openpyxl
import glob
from openpyxl import Workbook
import win32com.client
import re
import datetime as dt

path = r'C:\Test\Weekly review form'

wb = Workbook()

#return to the home directory
def home_dir():
    os.chdir(path)
    return

#open excel record
def open_review_form():
    global wb
    for filename in glob.glob(os.path.join(path, '*.xlsx')):
        print(filename)    
        wb = openpyxl.open(filename)
    return

#open the intern tab in the excel record
def open_tab(tab_name):
    global ws
    ws = wb[tab_name]
    return

#list out all employee names (including both intern and account manager)
def employee_names(col_name):
    global employee_list
    global ws
    employee_list = []
    for col in ws[col_name]:
        employee_list.append(col.value)
    employee_list.pop(0)
    return employee_list

def check_inbox(mail_box):
    outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')
    inbox = outlook.Folders(mail_box).Folders('Inbox')    
    messages = inbox.Items
    
    for i in messages:
        subject = i.Subject
        print(i.Subject)
    
    message = messages.GetLast()
    body_content = message.body
    print (body_content)

def timestamp():
    timestamp = dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return str(timestamp)

#open record file
home_dir()
open_review_form()

#read the intern 2021 tab
open_tab('intern 2021')
#ws['E3'].value
employee_names('B')

#read the consultant tab
open_tab('consultant')
employee_names('B')

#check Career UK email for form completion
check_inbox('Career UK')




xfile.save('text2.xlsx')






outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')
inbox = outlook.Folders('Career UK').Folders('Inbox')    

messages = inbox.Items

pattern = ['Weekly Intern Review Form', 'Weekly Report - Account Manager']


list_of_ppl_completed_form = []
num_of_ppl_completed_form = 0
for i in pattern:
    print('Looking for ' + str(i) + ' in inbox')
    
    for j, k in enumerate(messages):
        count = j + 1
        subject = k.Subject
        print('Parsing item ' + str(count))
                
        if re.search(i, subject):
            print(timestamp() + ': Found the following match: ' + str(count))
            num_of_ppl_completed_form =  num_of_ppl_completed_form + 1
            txt = re.compile('Name \* \\t([A-z][a-z]+ [A-Z][a-z]+ \([A-Z][a-z]+\)|[A-z][a-z]+ [A-Z][a-z]+)', re.IGNORECASE)
            list_of_ppl_completed_form.append(txt.findall(k.Body))
                        
        else:
            print(timestamp()+ ': No match for item ' + str(count))
        if count > 10:
            break
        
    if num_of_ppl_completed_form == len(list_of_ppl_completed_form):
        print('The number of people completed form matches with the length of the extracted list of people completed the form')
    else:
        print('ERROR: ' + str(num_of_ppl_completed_form) + ' completed the form but only ' + str(count) + ' people in the list.')

list_of_ppl_completed_form













pattern = re.compile('(Weekly Intern Review Form)|(Weekly Report - Account Manager)', re.IGNORECASE)


list_of_ppl_completed_form = []
num_of_ppl_completed_form = 0

for j, k in enumerate(messages):
    count = j + 1
    subject = k.Subject
    print('Parsing item ' + str(count))
            
    if re.search(i, subject):
        print(timestamp() + ': Found the following match: ' + str(count))
        num_of_ppl_completed_form =  num_of_ppl_completed_form + 1
        txt = re.compile('Name \* \\t([A-z][a-z]+ [A-Z][a-z]+ \([A-Z][a-z]+\)|[A-z][a-z]+ [A-Z][a-z]+)', re.IGNORECASE)
        list_of_ppl_completed_form.append(txt.findall(k.Body))
                    
    else:
        print(timestamp()+ ': No match for item ' + str(count))
    if count > 10:
        break
    
if num_of_ppl_completed_form == len(list_of_ppl_completed_form):
    print('The number of people completed form matches with the length of the extracted list of people completed the form')
else:
    print('ERROR: ' + str(num_of_ppl_completed_form) + ' completed the form but only ' + str(count) + ' people in the list.')


found = False
for match in pattern.findall(messages[6].Subject):
    if txt.findall(messages[6].Body):
        list_of_ppl_completed_form.append(txt.findall(messages[6].Body)[0])
        found = True
    if not found:
        list_of_ppl_completed_form.append(None)










